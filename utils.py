# expense-categorizer/utils.py
"""
Utility functions for expense categorization
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any


def load_transactions_from_file(filepath: Path) -> List[str]:
    """
    Load transaction descriptions from various file formats

    Args:
        filepath: Path to input file (JSON, TXT, CSV)

    Returns:
        List of transaction descriptions
    """
    descriptions = []

    if filepath.suffix == '.json':
        with open(filepath, 'r') as f:
            data = json.load(f)

        if isinstance(data, list):
            for item in data:
                if isinstance(item, str):
                    descriptions.append(item)
                elif isinstance(item, dict):
                    # Try common field names
                    desc = item.get('description') or item.get('Description') or \
                           item.get('desc') or item.get('text') or \
                           item.get('transaction') or item.get('memo')
                    if desc:
                        descriptions.append(str(desc))

    elif filepath.suffix == '.csv':
        with open(filepath, 'r', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Try common column names
                desc = row.get('description') or row.get('Description') or \
                       row.get('desc') or row.get('text') or \
                       row.get('transaction') or row.get('memo')
                if desc:
                    descriptions.append(str(desc))

    elif filepath.suffix == '.txt':
        with open(filepath, 'r') as f:
            descriptions = [line.strip() for line in f if line.strip()]

    return descriptions


def generate_statistics(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Generate statistics from classification results

    Args:
        results: List of classification results

    Returns:
        Dictionary containing statistics
    """
    total = len(results)
    if total == 0:
        return {}

    # Category counts
    category_counts = {}
    subcategory_counts = {}
    confidence_sum = 0

    for r in results:
        cat = r.get('category', 'unknown')
        category_counts[cat] = category_counts.get(cat, 0) + 1

        subcat = r.get('subcategory')
        if subcat:
            subcategory_counts[subcat] = subcategory_counts.get(subcat, 0) + 1

        confidence_sum += r.get('confidence', 0)

    # Sort by count
    sorted_categories = sorted(
        category_counts.items(),
        key=lambda x: x[1],
        reverse=True
    )

    sorted_subcategories = sorted(
        subcategory_counts.items(),
        key=lambda x: x[1],
        reverse=True
    )

    stats = {
        'total_transactions': total,
        'average_confidence': confidence_sum / total if total > 0 else 0,
        'categories': {
            cat: {
                'count': count,
                'percentage': (count / total) * 100
            }
            for cat, count in sorted_categories
        },
        'subcategories': {
            subcat: {
                'count': count,
                'percentage': (count / total) * 100
            }
            for subcat, count in sorted_subcategories
        } if sorted_subcategories else {}
    }

    return stats


def export_to_excel(results: List[Dict[str, Any]], filepath: Path):
    """
    Export results to Excel format (requires openpyxl)

    Args:
        results: List of classification results
        filepath: Output Excel file path
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Classifications"

        # Headers
        headers = ['Description', 'Category', 'Subcategory', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.get('description', ''))
            ws.cell(row=row, column=2, value=result.get('category', ''))
            ws.cell(row=row, column=3, value=result.get('subcategory', ''))
            ws.cell(row=row, column=4, value=result.get('confidence', 0))

        # Add statistics sheet
        stats = generate_statistics(results)
        ws2 = wb.create_sheet("Statistics")

        ws2.cell(row=1, column=1, value="Total Transactions")
        ws2.cell(row=1, column=2, value=stats.get('total_transactions', 0))

        ws2.cell(row=2, column=1, value="Average Confidence")
        ws2.cell(row=2, column=2, value=f"{stats.get('average_confidence', 0):.2%}")

        # Category breakdown
        ws2.cell(row=4, column=1, value="Category")
        ws2.cell(row=4, column=2, value="Count")
        ws2.cell(row=4, column=3, value="Percentage")

        row = 5
        for cat, data in stats['categories'].items():
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=data['count'])
            ws2.cell(row=row, column=3, value=f"{data['percentage']:.1f}%")
            row += 1

        # Adjust column widths
        for ws in [wb.active, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return True

    except ImportError:
        return False


def validate_categories(categories: List[str]) -> List[str]:
    """
    Validate and clean category names

    Args:
        categories: List of category names

    Returns:
        Cleaned list of categories
    """
    cleaned = []
    seen = set()

    for cat in categories:
        # Clean and normalize
        cat = cat.strip().lower()
        if cat and cat not in seen:
            cleaned.append(cat)
            seen.add(cat)

    # Ensure at least one category
    if not cleaned:
        cleaned = ['others']

    return cleaned